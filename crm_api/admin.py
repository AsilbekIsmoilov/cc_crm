from datetime import datetime
from io import StringIO
from django import forms
from django.contrib import admin, messages
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import path
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import *
from crm_api.services.users import bulk_create_operators, build_csv_from_results

def _current_changelist_queryset(modeladmin, request):
    ChangeList = modeladmin.get_changelist(request)
    cl = ChangeList(
        request, modeladmin.model, modeladmin.list_display, modeladmin.list_display_links,
        modeladmin.list_filter, modeladmin.date_hierarchy, modeladmin.search_fields,
        modeladmin.list_select_related, modeladmin.list_per_page, modeladmin.list_max_show_all,
        modeladmin.list_editable, modeladmin,modeladmin.sortable_by,modeladmin.search_help_text,
    )
    return cl.get_queryset(request)

def _export_status_calls_xlsx(qs):
    allowed = {choice for choice, _ in STATUS_CALL_CHOICES}
    qs = qs.filter(status_call__in=allowed).exclude(status_call__isnull=True).exclude(status_call="")

    wb = Workbook()
    ws = wb.active
    ws.title = "STATUS_CALL_CHOICES"

    columns = [
        ("MSISDN", "msisdn"),
        ("Клиент", "client"),
        ("Статус", "status"),
        ("Филиал", "branches"),
        ("Тариф", "rate_plan"),
        ("Абонплата", "subscription_fee"),
        ("Баланс", "balance"),
        ("Статус звонка", "status_call"),
        ("Результат звонка", "call_result"),
        ("Ответ абонента", "abonent_answer"),
        ("Технология", "tech"),
        ("Обновлён", "updated_at"),
        ("Кто звонил", "fixed_by"),
        ("Примечание", "note"),
    ]

    ws.append([c[0] for c in columns])

    for obj in qs.iterator():
        row = []
        for _, attr in columns:
            value = getattr(obj, attr, "")
            if attr == "fixed_by":
                value = obj.who_called
            if isinstance(value, datetime):
                value = timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S")
            row.append("" if value is None else value)
        ws.append(row)

    for idx, (header, _) in enumerate(columns, start=1):
        max_len = len(header)
        for cell_val in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, values_only=True):
            for v in cell_val:
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    now = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
    resp['Content-Disposition'] = f'attachment; filename="status_call_report_{now}.xlsx"'
    wb.save(resp)
    return resp


BASE_LIST_DISPLAY = (
    "msisdn", "client", "status", "branches", "rate_plan",
    "subscription_fee", "balance",
    "status_call", "call_result", "abonent_answer", "tech",
    "updated_at", 'fixed_by'
)

BASE_FIELDSETS = (
    ("Идентификация", {"fields": ("msisdn", "phone", "client", "account")}),
    ("Организационное", {"fields": ("departments", "branches")}),
    ("Тариф/баланс", {"fields": ("rate_plan", "subscription_fee", "balance")}),
    ("Статус абонента", {"fields": ("status", "status_from", "days_in_status", "write_offs_date")}),
    ("Фиксация", {"fields": ("status_call", "call_result", "abonent_answer", "note", "tech")}),
    ("Служебные", {"fields": ("created_at", "updated_at")}),
)

@admin.register(Actives)
class ActivesAdmin(admin.ModelAdmin):
    list_display = BASE_LIST_DISPLAY
    list_display_links = ("msisdn", "client")
    list_editable = ("status_call", "call_result", "abonent_answer", "tech")
    search_fields = ("msisdn", "phone", "client", "account", "note")
    list_filter = (
        ("created_at", admin.DateFieldListFilter),
        ("updated_at", admin.DateFieldListFilter),
        "branches", "rate_plan", "status",
        "status_call", "call_result", "abonent_answer", "tech",
        "fixed_by",
    )
    readonly_fields = ("created_at", "updated_at")
    fieldsets = BASE_FIELDSETS
    ordering = ("-updated_at",)
    list_per_page = 50
    date_hierarchy = "created_at"
    save_on_top = True
    empty_value_display = "—"

    change_list_template = "admin/crm_api/actives/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        my = [
            path(
                "export_status_call/",
                self.admin_site.admin_view(self.export_status_call_view),
                name="crm_api_actives_export_status_call",
            ),
        ]
        return my + urls

    def export_status_call_view(self, request):
        if not self.has_view_permission(request):
            raise PermissionDenied
        qs = _current_changelist_queryset(self, request)
        return _export_status_calls_xlsx(qs)

    def save_model(self, request, obj, form, change):
        fixation_fields = {"status_call", "call_result", "abonent_answer", "note", "tech"}
        changed = set(form.changed_data or [])
        if change and changed.intersection(fixation_fields):
            obj.fixed_by = request.user
            obj.fixed_at = timezone.now()
        super().save_model(request, obj, form, change)


@admin.register(Suspends)
class SuspendsAdmin(ActivesAdmin):
    change_list_template = "admin/crm_api/suspends/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        my = [
            path(
                "export_status_call/",
                self.admin_site.admin_view(self.export_status_call_view),
                name="crm_api_suspends_export_status_call",
            ),
        ]
        return my + urls

    def export_status_call_view(self, request):
        if not self.has_view_permission(request):
            raise PermissionDenied
        qs = _current_changelist_queryset(self, request)
        return _export_status_calls_xlsx(qs)



class BulkOperatorsForm(forms.Form):
    count = forms.IntegerField(min_value=1, max_value=5000, initial=50, label="Сколько создать")
    prefix = forms.CharField(max_length=50, initial="operator", label="Префикс логина")
    start = forms.IntegerField(min_value=0, initial=0, label="Начать с индекса")
    reset_existing = forms.BooleanField(required=False, initial=False, label="Сбросить пароли существующим")


User = get_user_model()


@admin.register(User)
class UserAdmin(DjangoUserAdmin):
    # ссылка в тулбаре списка пользователей
    change_list_template = "admin/crm_api/user/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        my = [
            path(
                "bulk-create-operators/",
                self.admin_site.admin_view(self.bulk_create_operators_view),
                name="crm_api_user_bulk_create_operators",
            ),
        ]
        return my + urls

    def bulk_create_operators_view(self, request):
        """
        GET — форма параметров.
        POST — создаём пользователей и отдаём CSV на скачивание.
        """
        initial = {"count": 50, "prefix": "operator", "start": 0, "reset_existing": False}
        form = BulkOperatorsForm(request.POST or None, initial=initial)

        if request.method == "POST":
            if not self.has_change_permission(request):
                raise PermissionDenied

            data = form.cleaned_data if form.is_valid() else initial
            rows = bulk_create_operators(
                count=data["count"],
                prefix=data["prefix"],
                start=data["start"],
                reset_existing=data["reset_existing"],
            )
            csv_io: StringIO = build_csv_from_results(rows)
            filename = f"operators_{data['prefix']}_{int(timezone.now().timestamp())}.csv"
            resp = HttpResponse(csv_io.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp

        context = {
            **self.admin_site.each_context(request),
            "title": "Bulk create operators (download CSV)",
            "opts": User._meta,
            "form": form,
        }
        return render(request, "admin/crm_api/user/bulk_create_operators.html", context)




@admin.register(SBMSAccount)
class SBMSAccountAdmin(admin.ModelAdmin):
    list_display = ("label", "username", "is_active", "max_google_accounts", "google_accounts_count")
    list_filter = ("is_active",)
    search_fields = ("label", "username")


@admin.register(GoogleAccount)
class GoogleAccountAdmin(admin.ModelAdmin):
    list_display = ("label", "sbms_account", "google_email", "is_active")
    list_filter = ("is_active", "sbms_account")
    search_fields = ("label", "google_email", "user_data_dir")


admin.site.register(ExcelUpload)
admin.site.register(SbmsAudit)
admin.site.register(RecheckRun)


