import io
import threading


from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings
from django.utils import timezone
from django.db.models import Q
import sys
from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import DjangoModelPermissions, IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView
import csv
from rest_framework_simplejwt.tokens import RefreshToken, UntypedToken
from rest_framework_simplejwt.views import TokenVerifyView
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from rest_framework_simplejwt.backends import TokenBackend
from django.db.models.functions import TruncDate
from .services import *
from .models import *
from .serializers import *
import openpyxl
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from datetime import datetime, date, time, timedelta
from rest_framework.decorators import api_view
from django.db.models import Count

from .services.excel_importer import run_import

ORDERABLE = {
    "id", "created_at", "updated_at", "msisdn", "client", "rate_plan",
    "branches", "status", "subscription_fee", "balance", "status_call",
    "call_result", "abonent_answer", "tech", "fixed_at", "account", "phone",
}

ALLOWED_SEARCH_FIELDS = {
    "msisdn": "msisdn__icontains",
    "phone": "phone__icontains",
    "client": "client__icontains",
    "account": "account__icontains",
}


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 500


def _change_perm_code():
    return f"{Actives._meta.app_label}.change_{Actives._meta.model_name}"


def _parse_search_fields(request) -> list[str]:
    raw = request.query_params.getlist("fields") or request.query_params.get("fields")
    if isinstance(raw, str):
        fields = [p.strip() for p in raw.split(",")]
    else:
        fields = list(raw or [])
    seen = set()
    cleaned = []
    for f in fields:
        if f in ALLOWED_SEARCH_FIELDS and f not in seen:
            cleaned.append(f)
            seen.add(f)
    return cleaned


def _apply_filters(request, qs):
    q = (request.query_params.get("q") or "").strip()
    fields = _parse_search_fields(request)

    if q:
        if fields:
            cond = Q()
            for f in fields:
                cond |= Q(**{ALLOWED_SEARCH_FIELDS[f]: q})
            qs = qs.filter(cond)
        else:
            cond = Q()
            for lookup in ALLOWED_SEARCH_FIELDS.values():
                cond |= Q(**{lookup: q})
            qs = qs.filter(cond)

    ordering = request.query_params.get("ordering")
    if ordering:
        fld = ordering.lstrip("-")
        if fld in ORDERABLE:
            qs = qs.order_by(ordering)
    return qs



class ActivesViewSet(viewsets.ModelViewSet):
    queryset = Actives.objects.all().order_by("-created_at")
    serializer_class = ActivesSerializer
    permission_classes = [permissions.IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        qs = super().get_queryset()
        status_filter = self.request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status__icontains=status_filter)
        return _apply_filters(self.request, qs)

    @action(detail=True, methods=["get", "patch"], url_path="fixation")
    def fixation(self, request, pk=None):
        abonent = self.get_object()

        if request.method == "GET":
            return Response(ActivesSerializer(abonent).data, status=status.HTTP_200_OK)

        perm_code = _change_perm_code()
        if not (request.user.is_superuser or request.user.has_perm(perm_code)):
            raise PermissionDenied(f"Недостаточно прав: требуется '{perm_code}'.")

        ser = ActivesFixationWriteSerializer(abonent, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()

        abonent.fixed_by = request.user
        abonent.fixed_at = timezone.now()
        abonent.save(update_fields=["fixed_by", "fixed_at", "updated_at"])
        return Response(ActivesSerializer(abonent).data, status=status.HTTP_200_OK)


class SuspendsViewSet(viewsets.ModelViewSet):
    queryset = Suspends.objects.all().order_by("-created_at")
    serializer_class = ActivesSerializer
    permission_classes = [permissions.IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        qs = super().get_queryset()
        return _apply_filters(self.request, qs)

    @action(detail=True, methods=["get", "patch"], url_path="fixation")
    def fixation(self, request, pk=None):
        abonent = self.get_object()

        if request.method == "GET":
            return Response(ActivesSerializer(abonent).data, status=status.HTTP_200_OK)

        perm_code = _change_perm_code()
        if not (request.user.is_superuser or request.user.has_perm(perm_code)):
            raise PermissionDenied(f"Недостаточно прав: требуется '{perm_code}'.")

        ser = ActivesFixationWriteSerializer(abonent, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)

        with transaction.atomic():
            for f, v in ser.validated_data.items():
                setattr(abonent, f, v)
            abonent.fixed_by = request.user
            abonent.fixed_at = timezone.now()
            abonent.save()

            fixed = Fixeds.objects.create(
                msisdn=abonent.msisdn,
                departments=abonent.departments,
                status_from=abonent.status_from,
                days_in_status=abonent.days_in_status,
                write_offs_date=abonent.write_offs_date,
                client=abonent.client,
                rate_plan=abonent.rate_plan,
                balance=abonent.balance,
                subscription_fee=abonent.subscription_fee,
                account=abonent.account,
                branches=abonent.branches,
                status=abonent.status,
                phone=abonent.phone,
                status_call=abonent.status_call,
                call_result=abonent.call_result,
                abonent_answer=abonent.abonent_answer,
                note=abonent.note,
                tech=abonent.tech,
                fixed_by=abonent.fixed_by,
                fixed_at=abonent.fixed_at,
                created_at=abonent.created_at,
                updated_at=abonent.updated_at,
            )

            abonent.delete()

        return Response(FixedsSerializer(fixed).data, status=status.HTTP_201_CREATED)


class ExcelUploadViewSet(viewsets.ModelViewSet):
    queryset = ExcelUpload.objects.order_by("-uploaded_at")
    serializer_class = ExcelUploadSerializer
    permission_classes = [permissions.IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination


class OperatorsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.filter(role=User.ROLE_OPERATOR).order_by("fio", "username")
    serializer_class = OperatorSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None



class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return Response(MeSerializer(request.user).data, status=status.HTTP_200_OK)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        refresh = request.data.get("refresh")
        if not refresh:
            return Response({"detail": "Field 'refresh' is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            token = RefreshToken(refresh)
            token.blacklist()
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({"detail": "Logged out."}, status=status.HTTP_205_RESET_CONTENT)


class MyTokenVerifyView(TokenVerifyView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        token_str = request.data.get("token")
        if not token_str:
            return Response({"detail": "Field 'token' is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            UntypedToken(token_str)
        except (InvalidToken, TokenError):
            return Response({"detail": "Token is invalid or expired.", "code": "token_not_valid"}, status=401)

        backend = TokenBackend(
            algorithm=getattr(settings, "SIMPLE_JWT", {}).get("ALGORITHM", "HS256"),
            signing_key=getattr(settings, "SIMPLE_JWT", {}).get("SIGNING_KEY", settings.SECRET_KEY),
            verifying_key=getattr(settings, "SIMPLE_JWT", {}).get("VERIFYING_KEY", None),
            audience=getattr(settings, "SIMPLE_JWT", {}).get("AUDIENCE", None),
            issuer=getattr(settings, "SIMPLE_JWT", {}).get("ISSUER", None),
            jwk_url=getattr(settings, "SIMPLE_JWT", {}).get("JWK_URL", None),
            leeway=getattr(settings, "SIMPLE_JWT", {}).get("LEEWAY", 0),
        )
        payload = backend.decode(token_str, verify=True)

        user_info = None
        uid = payload.get("user_id") or payload.get("user") or payload.get("sub")
        if uid:
            try:
                u = User.objects.only("id", "username", "first_name", "last_name", "email").get(pk=uid)
                user_info = MeSerializer(u).data
            except User.DoesNotExist:
                user_info = None

        return Response({"valid": True, "payload": payload, "user": user_info}, status=status.HTTP_200_OK)

def _fmt_local(dt, fmt="%Y-%m-%d %H:%M:%S") -> str:
    if not dt:
        return ""
    try:
        # If dt is naïve, attach default timezone
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_default_timezone())
        # Convert to localtime and format
        return timezone.localtime(dt).strftime(fmt)
    except Exception:
        # Fallback: best-effort formatting without tz ops
        try:
            return dt.strftime(fmt)
        except Exception:
            return ""

def export_all_suspends(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suspends"

    ws.append([
        "DEPARTMENTS", "MSISDN", "Статус","CLIENT", "PHONE", "BRANCHES","Дата с которой Статус","[Дней в статусе]","Дата Списания АП","RATE_PLAN","Баланс","Абон плата","ACCOUNT",
        "Статус звонка", "Результат обзвона", "Ответ абонента",
        "Дата обзвона","Оператор"
    ])

    for obj in Suspends.objects.all():
        created_at_str = _fmt_local(obj.created_at)
        fixed_at_str   = _fmt_local(obj.fixed_at)

        ws.append([
            obj.departments,
            obj.msisdn,
            obj.status,
            obj.client,
            obj.phone,
            obj.branches,
            obj.status_from,
            obj.days_in_status,
            obj.write_offs_date,
            obj.rate_plan,
            obj.balance,
            obj.subscription_fee,
            obj.account,
            obj.status_call,
            obj.call_result,
            obj.abonent_answer,
            fixed_at_str,
            obj.who_called,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="all_suspends.xlsx"'
    wb.save(response)
    return response

def export_all_actives(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actives"

    ws.append([
        "DEPARTMENTS", "MSISDN", "Статус","CLIENT", "PHONE", "BRANCHES","Дата с которой Статус","[Дней в статусе]","Дата Списания АП","RATE_PLAN","Баланс","Абон плата","ACCOUNT",
        "Статус звонка", "Результат обзвона", "Ответ абонента",
        "Дата обзвона"
    ])

    for obj in Actives.objects.all():
        created_at_str = _fmt_local(obj.created_at)
        fixed_at_str   = _fmt_local(obj.fixed_at)

        ws.append([
            obj.departments,
            obj.msisdn,
            obj.status,
            obj.client,
            obj.phone,
            obj.branches,
            obj.status_from,
            obj.days_in_status,
            obj.write_offs_date,
            obj.rate_plan,
            obj.balance,
            obj.subscription_fee,
            obj.account,
            obj.status_call,
            obj.call_result,
            obj.abonent_answer,
            fixed_at_str,
            obj.who_called,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="all_actives.xlsx"'
    wb.save(response)
    return response

def _capture_run(func):
    buffer = io.StringIO()
    old_stdout, old_stderr = sys.stdout, sys.stderr
    sys.stdout = sys.stderr = buffer
    try:
        result = func()
    except Exception as e:
        print("❌ Ошибка:", e)
        result = None
    finally:
        sys.stdout, sys.stderr = old_stdout, old_stderr

    log_text = buffer.getvalue()
    return log_text, result

@staff_member_required
def export_suspends_phones_csv(request):
    qs = (Suspends.objects
          .values_list("phone", flat=True)
          .exclude(phone__isnull=True)
          .exclude(phone__exact=""))

    if request.GET.get("distinct"):
        qs = qs.distinct()

    now = timezone.localtime()
    filename = f"suspends_phones_{now:%Y-%m-%d_%H-%M-%S}.csv"

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write("\ufeff")

    writer = csv.writer(resp, lineterminator="\n")
    writer.writerow(["PHONE"])

    for phone in qs.iterator(chunk_size=5000):
        p = (phone or "").strip()
        if p:
            writer.writerow([p])

    return resp

class ImportUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAdminUser]

    def post(self, request, *args, **kwargs):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Приложите файл в поле 'file'."},
                            status=status.HTTP_400_BAD_REQUEST)

        job = UploadJob.objects.create(
            created_by=request.user if request.user.is_authenticated else None,
            excel_file=f,
            status="pending",
        )
        t = threading.Thread(target=run_import, args=(job.id,), daemon=True)
        t.start()

        return Response({"job_id": job.id}, status=status.HTTP_201_CREATED)

class ImportStatusView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, job_id: int, *args, **kwargs):
        try:
            job = UploadJob.objects.get(id=job_id)
        except UploadJob.DoesNotExist:
            return Response({"detail": "Job not found."}, status=status.HTTP_404_NOT_FOUND)

        data = UploadJobSerializer(job).data
        return Response(data, status=status.HTTP_200_OK)

class FixedsViewSet(viewsets.ModelViewSet):
    queryset = Fixeds.objects.all()
    serializer_class = FixedsSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['client', 'account', 'msisdn', 'phone']
    search_param = 'q'
    ordering_fields = ['fixed_at','updated_at','created_at','msisdn','client','account','phone']
    ordering = ['-fixed_at']


class SearchSuspendsFixeds(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            page = max(int(request.query_params.get("page") or 1), 1)
        except Exception:
            page = 1
        try:
            page_size = min(max(int(request.query_params.get("page_size") or 50), 1), 500)
        except Exception:
            page_size = 50

        ordering = request.query_params.get("ordering") or "-created_at"
        fld = ordering.lstrip("-")
        if fld not in ORDERABLE:
            ordering = "-created_at"
            fld = "created_at"
        reverse = ordering.startswith("-")

        # --- исходные qs + фильтры ---
        qs_s_base = _apply_filters(request, Suspends.objects.all())
        qs_f_base = _apply_filters(request, Fixeds.objects.all())

        total = qs_s_base.count() + qs_f_base.count()
        need = page * page_size

        qs_s = qs_s_base.order_by(ordering)[:need]
        qs_f = qs_f_base.order_by(ordering)[:need]

        def sort_key(o):
            v = getattr(o, fld, None)
            return (v is None, v)

        merged = [("suspends", o) for o in qs_s] + [("fixeds", o) for o in qs_f]
        merged.sort(key=lambda t: sort_key(t[1]), reverse=reverse)

        start, end = (page - 1) * page_size, (page - 1) * page_size + page_size
        page_slice = merged[start:end]

        def norm(src, obj):
            return {
                "id": obj.id,
                "msisdn": obj.msisdn,
                "departments": obj.departments,
                "status_from": obj.status_from,
                "days_in_status": obj.days_in_status,
                "write_offs_date": obj.write_offs_date,
                "client": obj.client,
                "rate_plan": obj.rate_plan,
                "balance": obj.balance,
                "subscription_fee": obj.subscription_fee,
                "account": obj.account,
                "branches": obj.branches,
                "status": obj.status,
                "phone": obj.phone,
                "status_call": obj.status_call,
                "call_result": obj.call_result,
                "abonent_answer": obj.abonent_answer,
                "note": obj.note,
                "tech": obj.tech,
                "called_by_id": getattr(obj, "fixed_by_id", None),
                "called_by": obj.who_called if hasattr(obj, "who_called") else "",
                "called_at": getattr(obj, "fixed_at", None),
                "created_at": getattr(obj, "created_at", None),
                "updated_at": getattr(obj, "updated_at", None),
                "source": src,
            }

        results = [norm(src, obj) for src, obj in page_slice]
        return Response({"count": total, "next": None, "previous": None, "results": results}, status=status.HTTP_200_OK)





def _fmt(dt):
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def _day_range(d: date):
    start = datetime.combine(d, time.min)
    end = start + timedelta(days=1)
    return start, end

def _month_range(y: int, m: int):
    first = datetime(y, m, 1)
    if m == 12:
        nxt = datetime(y + 1, 1, 1)
    else:
        nxt = datetime(y, m + 1, 1)
    return first, nxt

def _write_fixeds_sheet(ws, qs):
    ws.title = "Fixeds"
    ws.append([
        "Департамент", "MSISDN", "Статус", "Клиент", "Телефон", "Филиал",
        "Дата с которой статус", "Дней в статусе", "Дата списания АП",
        "Тариф", "Баланс", "Абон плата", "Лицевой счёт",
        "Статус звонка", "Результат обзвона", "Ответ абонента",
        "Дата обзвона", "Оператор",
    ])

    for obj in qs:
        ws.append([
            obj.departments,
            obj.msisdn,
            obj.status,
            obj.client,
            obj.phone,
            obj.branches,
            obj.status_from,
            obj.days_in_status,
            obj.write_offs_date,
            obj.rate_plan,
            obj.balance,
            obj.subscription_fee,
            obj.account,
            obj.status_call,
            obj.call_result,
            obj.abonent_answer,
            _fmt(obj.fixed_at),
            obj.who_called,
        ])

def _xlsx_response(wb, filename: str):
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp



def export_all_fixeds(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_fixeds_sheet(ws, Fixeds.objects.all().order_by("fixed_at", "id"))
    return _xlsx_response(wb, "fixeds_all.xlsx")


def export_fixeds_daily(request):
    date_str = request.GET.get("date", "")
    if date_str:
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Bad date format, use YYYY-MM-DD", status=400)
    else:
        d = date.today()

    start, end = _day_range(d)
    qs = (Fixeds.objects
          .filter(fixed_at__gte=start, fixed_at__lt=end)
          .exclude(fixed_at__isnull=True)
          .order_by("fixed_at", "id"))

    wb = openpyxl.Workbook()
    ws = wb.active
    _write_fixeds_sheet(ws, qs)
    return _xlsx_response(wb, f"fixeds_day_{d.strftime('%Y-%m-%d')}.xlsx")


def export_fixeds_monthly(request):
    month_str = request.GET.get("month", "")
    if month_str:
        try:
            y, m = map(int, month_str.split("-"))
        except Exception:
            return HttpResponse("Bad month format, use YYYY-MM", status=400)
    else:
        today = date.today()
        y, m = today.year, today.month

    start, end = _month_range(y, m)
    qs = (Fixeds.objects
          .filter(fixed_at__gte=start, fixed_at__lt=end)
          .exclude(fixed_at__isnull=True)
          .order_by("fixed_at", "id"))

    wb = openpyxl.Workbook()
    ws = wb.active
    _write_fixeds_sheet(ws, qs)
    return _xlsx_response(wb, f"fixeds_month_{y:04d}-{m:02d}.xlsx")


class MoveSuspendsToFixedsAPIView(APIView):
    permission_classes = [IsAdminUser]
    def post(self, request, *args, **kwargs):
        from .models import move_suspends_with_status_call_to_fixeds
        with transaction.atomic():
            result = move_suspends_with_status_call_to_fixeds()
        payload = {"detail": "Move job finished"}
        if isinstance(result, dict):
            payload.update(result)
        elif isinstance(result, (list, tuple)):
            payload["result"] = list(result)
        elif isinstance(result, int):
            payload["moved"] = result

        return Response(payload, status=status.HTTP_200_OK)


@api_view(["POST"])
def resolve_msisdn(request):
    """
    Принимает список номеров, возвращает соответствующие MSISDN.
    """
    numbers = request.data.get("numbers", [])
    result = []
    for num in numbers:
        # ищем запись в Suspends по номеру (phone)
        rec = Suspends.objects.filter(phone=num).first()
        if rec and rec.msisdn:
            result.append({"number": num, "msisdn": rec.msisdn})
        else:
            result.append({"number": num, "msisdn": None})
    return Response(result)


class OperatorStatisticsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        now = timezone.now()

        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())  # понедельник
        month_start = today_start.replace(day=1)

        qs = Fixeds.objects.filter(fixed_by=user)

        stats = {
            "today": qs.filter(fixed_at__gte=today_start).count(),
            "week": qs.filter(fixed_at__gte=week_start).count(),
            "month": qs.filter(fixed_at__gte=month_start).count(),
            "total": qs.count(),
        }

        status_stats = (
            qs.values("status_call")
              .annotate(count=Count("id"))
              .order_by()
        )

        status_dict = {row["status_call"] or "не указано": row["count"] for row in status_stats}

        return Response({
            "general": stats,
            "by_status_call": status_dict
        }, status=status.HTTP_200_OK)



class OperatorDailyStatsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        qs = (
            Fixeds.objects
            .filter(fixed_by=user, fixed_at__gte=month_start)
            .annotate(day=TruncDate("fixed_at"))
            .values("day")
            .annotate(count=Count("id"))
            .order_by("day")
        )

        data = [
            {"date": row["day"].strftime("%Y-%m-%d"), "count": row["count"]}
            for row in qs
        ]

        return Response(data, status=status.HTTP_200_OK)